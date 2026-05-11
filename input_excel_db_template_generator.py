import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_exam_scheduler_template():
    """Créer un modèle Excel vide avec le format requis pour le planificateur d'examens"""
    
    # Créer un nouveau classeur
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # Définir les styles
    header_font = Font(name='Arial', bold=True, size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # =============== FEUILLE 1: Salles ===============
    rooms_sheet = wb.create_sheet("Salles")
    
    # En-têtes pour la feuille Salles
    rooms_headers = ["Nom de la Salle", "Capacité"]
    rooms_sheet.append(rooms_headers)
    
    # Styliser les en-têtes
    for col_num, header in enumerate(rooms_headers, 1):
        cell = rooms_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Définir les largeurs de colonnes
    rooms_sheet.column_dimensions['A'].width = 20  # Nom de la Salle
    rooms_sheet.column_dimensions['B'].width = 12  # Capacité
    
    # =============== FEUILLE 2: Groupes ===============
    groups_sheet = wb.create_sheet("Groupes")
    
    # En-têtes pour la feuille Groupes (jusqu'à 5 matières par groupe)
    max_subjects = 5
    groups_headers = ["Nom du Groupe", "Nb Matières"]
    
    # Ajouter les colonnes Matière et Durée
    for i in range(1, max_subjects + 1):
        groups_headers.extend([f"Matière {i}", f"Durée {i}"])
    
    groups_sheet.append(groups_headers)
    
    # Styliser les en-têtes
    for col_num, header in enumerate(groups_headers, 1):
        cell = groups_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Définir les largeurs de colonnes
    groups_sheet.column_dimensions['A'].width = 18  # Nom du Groupe
    groups_sheet.column_dimensions['B'].width = 12  # Nb Matières
    
    # Colonnes Matière et Durée
    for i in range(3, len(groups_headers) + 1):
        col_letter = groups_sheet.cell(row=1, column=i).column_letter
        if i % 2 == 1:  # Colonnes Matière
            groups_sheet.column_dimensions[col_letter].width = 20
        else:  # Colonnes Durée
            groups_sheet.column_dimensions[col_letter].width = 12
    
    # =============== FEUILLE 3: Enseignants ===============
    teachers_sheet = wb.create_sheet("Enseignants")
    
    # En-têtes pour la feuille Enseignants (jusqu'à 4 matières par enseignant)
    max_teacher_subjects = 4
    teachers_headers = ["Nom de l'Enseignant", "Nb Matières"]
    
    for i in range(1, max_teacher_subjects + 1):
        teachers_headers.append(f"Matière {i}")
    
    teachers_sheet.append(teachers_headers)
    
    # Styliser les en-têtes
    for col_num, header in enumerate(teachers_headers, 1):
        cell = teachers_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Définir les largeurs de colonnes
    teachers_sheet.column_dimensions['A'].width = 22  # Nom de l'Enseignant
    teachers_sheet.column_dimensions['B'].width = 12  # Nb Matières
    
    # Colonnes Matière
    for i in range(3, len(teachers_headers) + 1):
        col_letter = teachers_sheet.cell(row=1, column=i).column_letter
        teachers_sheet.column_dimensions[col_letter].width = 20
    
    # =============== AJOUTER FEUILLE INSTRUCTIONS ===============
    instructions_sheet = wb.create_sheet("Instructions")
    
    # Instructions en français
    instructions = [
        ["📋 INSTRUCTIONS - MODE D'EMPLOI"],
        [""],
        ["🏫 FEUILLE 1 - Salles :"],
        ["• Nom de la Salle : Entrez l'identifiant de la salle (ex: 101, Amphi A)"],
        ["• Capacité : Nombre de groupes que cette salle peut accueillir simultanément"],
        [""],
        ["👥 FEUILLE 2 - Groupes :"],
        ["• Nom du Groupe : Entrez l'identifiant du groupe (ex: 1ère AP1, Master 1)"],
        ["• Nb Matières : Nombre total de matières pour ce groupe"],
        ["• Matière 1, 2, etc. : Entrez les noms des matières"],
        ["• Durée 1, 2, etc. : Entrez la durée d'examen en minutes"],
        ["• Laissez vides les colonnes Matière/Durée non utilisées"],
        [""],
        ["👨‍🏫 FEUILLE 3 - Enseignants :"],
        ["• Nom de l'Enseignant : Entrez le nom complet de l'enseignant"],
        ["• Nb Matières : Nombre de matières enseignées par cet enseignant"],
        ["• Matière 1, 2, etc. : Entrez les matières enseignées"],
        ["• Les enseignants NE PEUVENT PAS surveiller leurs propres matières"],
        ["• Laissez vides les colonnes Matière non utilisées"],
        [""],
        ["⚠️ NOTES IMPORTANTES :"],
        ["• Supprimez cette feuille Instructions avant d'exécuter le planificateur"],
        ["• Toutes les durées doivent être en minutes (ex: 90, 120)"],
        ["• Capacité salle : 1 = groupe unique, 2+ = groupes multiples"],
        ["• Matières enseignant : Utilisées pour éviter l'auto-surveillance"],
        ["• Gardez les noms de feuilles exactement : Salles, Groupes, Enseignants"]
    ]
    
    # Ajouter les instructions
    for row_num, instruction in enumerate(instructions, 1):
        instructions_sheet.cell(row=row_num, column=1, value=instruction[0])
        
        # Styliser les en-têtes principaux
        if "INSTRUCTIONS" in instruction[0] or "FEUILLE" in instruction[0] or "NOTES IMPORTANTES" in instruction[0]:
            cell = instructions_sheet.cell(row=row_num, column=1)
            cell.font = Font(bold=True, size=12, color="1f4e79")
    
    # Définir la largeur de colonne pour les instructions
    instructions_sheet.column_dimensions['A'].width = 85
    
    # Déplacer la feuille Instructions en première position
    wb.move_sheet(instructions_sheet, 0)
    
    # Sauvegarder le modèle
    filename = "Modele_Planificateur_Examens.xlsx"
    wb.save(filename)
    
    print(f"✅ Modèle créé avec succès : {filename}")
    print(f"📁 Emplacement : {os.path.abspath(filename)}")
    print("\n📋 Le modèle inclut :")
    print("  • Feuille Instructions (à supprimer avant utilisation)")
    print("  • Feuille Salles (prête pour les données)")
    print("  • Feuille Groupes (prête pour les données)")  
    print("  • Feuille Enseignants (prête pour les données)")
    print("\n🚀 Prêt à utiliser avec votre planificateur d'examens !")

if __name__ == "__main__":
    create_exam_scheduler_template()