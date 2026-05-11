import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import pandas as pd
import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class GenerateurCalendrierExamens:
    def __init__(self, root):
        self.root = root
        self.root.title("Générateur des exemples de Calendrier d'Examens")
        self.root.geometry("800x600")
        
        # Obtenir le répertoire de l'exécutable
        if getattr(sys, 'frozen', False):
            # Si l'application est compilée avec PyInstaller
            self.exe_dir = os.path.dirname(sys.executable)
        else:
            # Si on exécute le script Python directement
            self.exe_dir = os.path.dirname(os.path.abspath(__file__))
        
        self.examens = []
        self.setup_ui()
        
    def setup_ui(self):
        # Style
        style = ttk.Style()
        style.theme_use('clam')
        
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configuration des poids pour le redimensionnement
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Titre
        titre = ttk.Label(main_frame, text="Générateur des exemples de Calendrier d'Examens", 
                         font=('Arial', 16, 'bold'))
        titre.grid(row=0, column=0, columnspan=3, pady=10)
        
        # Section d'ajout d'examens
        section_frame = ttk.LabelFrame(main_frame, text="Ajouter un Examen", padding="10")
        section_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        section_frame.columnconfigure(1, weight=1)
        
        # Nom de l'examen
        ttk.Label(section_frame, text="Nom de l'examen:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.nom_entry = ttk.Entry(section_frame, width=40)
        self.nom_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(5, 0), pady=2)
        
        # Matière
        ttk.Label(section_frame, text="Matière:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.matiere_entry = ttk.Entry(section_frame, width=40)
        self.matiere_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 0), pady=2)
        
        # Date
        ttk.Label(section_frame, text="Date (JJ/MM/AAAA):").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.date_entry = ttk.Entry(section_frame, width=40)
        self.date_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(5, 0), pady=2)
        
        # Heure de début
        ttk.Label(section_frame, text="Heure de début (HH:MM):").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.heure_debut_entry = ttk.Entry(section_frame, width=40)
        self.heure_debut_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(5, 0), pady=2)
        
        # Durée
        ttk.Label(section_frame, text="Durée (minutes):").grid(row=4, column=0, sticky=tk.W, pady=2)
        self.duree_entry = ttk.Entry(section_frame, width=40)
        self.duree_entry.grid(row=4, column=1, sticky=(tk.W, tk.E), padx=(5, 0), pady=2)
        
        # Lieu
        ttk.Label(section_frame, text="Lieu:").grid(row=5, column=0, sticky=tk.W, pady=2)
        self.lieu_entry = ttk.Entry(section_frame, width=40)
        self.lieu_entry.grid(row=5, column=1, sticky=(tk.W, tk.E), padx=(5, 0), pady=2)
        
        # Notes
        ttk.Label(section_frame, text="Notes:").grid(row=6, column=0, sticky=tk.W, pady=2)
        self.notes_entry = ttk.Entry(section_frame, width=40)
        self.notes_entry.grid(row=6, column=1, sticky=(tk.W, tk.E), padx=(5, 0), pady=2)
        
        # Bouton d'ajout
        btn_ajouter = ttk.Button(section_frame, text="Ajouter l'Examen", 
                               command=self.ajouter_examen)
        btn_ajouter.grid(row=7, column=0, columnspan=2, pady=10)
        
        # Liste des examens
        liste_frame = ttk.LabelFrame(main_frame, text="Liste des Examens", padding="10")
        liste_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        liste_frame.columnconfigure(0, weight=1)
        liste_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # Treeview pour afficher les examens
        columns = ("Nom", "Matière", "Date", "Heure", "Durée", "Lieu")
        self.tree = ttk.Treeview(liste_frame, columns=columns, show='headings', height=8)
        
        # Configuration des colonnes
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(liste_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Boutons d'action
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.grid(row=3, column=0, columnspan=3, pady=10)
        
        btn_supprimer = ttk.Button(buttons_frame, text="Supprimer l'Examen Sélectionné", 
                                 command=self.supprimer_examen)
        btn_supprimer.pack(side=tk.LEFT, padx=5)
        
        btn_modifier = ttk.Button(buttons_frame, text="Modifier l'Examen Sélectionné", 
                                command=self.modifier_examen)
        btn_modifier.pack(side=tk.LEFT, padx=5)
        
        btn_generer = ttk.Button(buttons_frame, text="Générer le Calendrier Excel", 
                               command=self.generer_calendrier_excel)
        btn_generer.pack(side=tk.LEFT, padx=5)
        
        btn_importer = ttk.Button(buttons_frame, text="Importer depuis Excel", 
                                command=self.importer_excel)
        btn_importer.pack(side=tk.LEFT, padx=5)
        
        btn_template = ttk.Button(buttons_frame, text="Créer Modèle de Planification", 
                                command=self.creer_modele_planification)
        btn_template.pack(side=tk.LEFT, padx=5)
        
        # Répertoire de sortie
        repertoire_frame = ttk.Frame(main_frame)
        repertoire_frame.grid(row=4, column=0, columnspan=3, pady=10, sticky=(tk.W, tk.E))
        repertoire_frame.columnconfigure(1, weight=1)
        
        ttk.Label(repertoire_frame, text="Répertoire de sortie:").grid(row=0, column=0, sticky=tk.W)
        self.repertoire_label = ttk.Label(repertoire_frame, text=self.exe_dir, 
                                        background='white', relief='sunken')
        self.repertoire_label.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
        
    def ajouter_examen(self):
        try:
            nom = self.nom_entry.get().strip()
            matiere = self.matiere_entry.get().strip()
            date_str = self.date_entry.get().strip()
            heure_debut = self.heure_debut_entry.get().strip()
            duree = self.duree_entry.get().strip()
            lieu = self.lieu_entry.get().strip()
            notes = self.notes_entry.get().strip()
            
            if not all([nom, matiere, date_str, heure_debut, duree]):
                messagebox.showerror("Erreur", "Veuillez remplir tous les champs obligatoires.")
                return
            
            # Validation de la date
            try:
                date_obj = datetime.strptime(date_str, "%d/%m/%Y")
            except ValueError:
                messagebox.showerror("Erreur", "Format de date invalide. Utilisez JJ/MM/AAAA.")
                return
            
            # Validation de l'heure
            try:
                heure_obj = datetime.strptime(heure_debut, "%H:%M")
            except ValueError:
                messagebox.showerror("Erreur", "Format d'heure invalide. Utilisez HH:MM.")
                return
            
            # Validation de la durée
            try:
                duree_int = int(duree)
                if duree_int <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Erreur", "La durée doit être un nombre positif de minutes.")
                return
            
            examen = {
                'nom': nom,
                'matiere': matiere,
                'date': date_obj,
                'heure_debut': heure_obj.time(),
                'duree': duree_int,
                'lieu': lieu,
                'notes': notes
            }
            
            self.examens.append(examen)
            self.mettre_a_jour_liste()
            self.vider_champs()
            messagebox.showinfo("Succès", "Examen ajouté avec succès!")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'ajout de l'examen: {str(e)}")
    
    def vider_champs(self):
        self.nom_entry.delete(0, tk.END)
        self.matiere_entry.delete(0, tk.END)
        self.date_entry.delete(0, tk.END)
        self.heure_debut_entry.delete(0, tk.END)
        self.duree_entry.delete(0, tk.END)
        self.lieu_entry.delete(0, tk.END)
        self.notes_entry.delete(0, tk.END)
    
    def mettre_a_jour_liste(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for examen in self.examens:
            self.tree.insert('', 'end', values=(
                examen['nom'],
                examen['matiere'],
                examen['date'].strftime("%d/%m/%Y"),
                examen['heure_debut'].strftime("%H:%M"),
                f"{examen['duree']} min",
                examen['lieu']
            ))
    
    def supprimer_examen(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un examen à supprimer.")
            return
        
        if messagebox.askyesno("Confirmation", "Êtes-vous sûr de vouloir supprimer cet examen?"):
            index = self.tree.index(selection[0])
            del self.examens[index]
            self.mettre_a_jour_liste()
            messagebox.showinfo("Succès", "Examen supprimé avec succès!")
    
    def modifier_examen(self):
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un examen à modifier.")
            return
        
        index = self.tree.index(selection[0])
        examen = self.examens[index]
        
        # Remplir les champs avec les données de l'examen sélectionné
        self.nom_entry.delete(0, tk.END)
        self.nom_entry.insert(0, examen['nom'])
        
        self.matiere_entry.delete(0, tk.END)
        self.matiere_entry.insert(0, examen['matiere'])
        
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, examen['date'].strftime("%d/%m/%Y"))
        
        self.heure_debut_entry.delete(0, tk.END)
        self.heure_debut_entry.insert(0, examen['heure_debut'].strftime("%H:%M"))
        
        self.duree_entry.delete(0, tk.END)
        self.duree_entry.insert(0, str(examen['duree']))
        
        self.lieu_entry.delete(0, tk.END)
        self.lieu_entry.insert(0, examen['lieu'])
        
        self.notes_entry.delete(0, tk.END)
        self.notes_entry.insert(0, examen['notes'])
        
        # Supprimer l'ancien examen
        del self.examens[index]
        self.mettre_a_jour_liste()
        
        messagebox.showinfo("Information", "Examen chargé pour modification. Cliquez sur 'Ajouter l'Examen' pour sauvegarder les modifications.")
    
    def generer_calendrier_excel(self):
        if not self.examens:
            messagebox.showwarning("Attention", "Aucun examen à exporter.")
            return
        
        try:
            # Préparer les données pour Excel
            donnees = []
            for examen in self.examens:
                heure_fin = (datetime.combine(datetime.today(), examen['heure_debut']) + 
                           timedelta(minutes=examen['duree'])).time()
                
                donnees.append({
                    'Nom de l\'Examen': examen['nom'],
                    'Matière': examen['matiere'],
                    'Date': examen['date'].strftime("%d/%m/%Y"),
                    'Jour de la Semaine': self.get_jour_semaine(examen['date']),
                    'Heure de Début': examen['heure_debut'].strftime("%H:%M"),
                    'Heure de Fin': heure_fin.strftime("%H:%M"),
                    'Durée (minutes)': examen['duree'],
                    'Lieu': examen['lieu'],
                    'Notes': examen['notes'],
                    'Temps Restant': self.calculer_temps_restant(examen['date'])
                })
            
            df = pd.DataFrame(donnees)
            
            # Trier par date
            df_sorted = df.sort_values('Date')
            
            # Créer le nom de fichier avec timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nom_fichier = f"calendrier_examens_{timestamp}.xlsx"
            chemin_fichier = os.path.join(self.exe_dir, nom_fichier)
            
            # Créer un fichier Excel avec mise en forme
            with pd.ExcelWriter(chemin_fichier, engine='openpyxl') as writer:
                df_sorted.to_excel(writer, sheet_name='Calendrier des Examens', index=False)
                
                # Obtenir la feuille de travail pour la mise en forme
                worksheet = writer.sheets['Calendrier des Examens']
                
                # Ajuster la largeur des colonnes
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Créer également une vue calendrier mensuelle
                self.creer_vue_calendrier_mensuel(writer, df_sorted)
            
            messagebox.showinfo("Succès", f"Calendrier généré avec succès!\nFichier sauvegardé: {chemin_fichier}")
            
            # Demander si l'utilisateur veut ouvrir le fichier
            if messagebox.askyesno("Ouvrir le fichier", "Voulez-vous ouvrir le fichier Excel maintenant?"):
                os.startfile(chemin_fichier)
                
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération du calendrier: {str(e)}")
    
    def creer_vue_calendrier_mensuel(self, writer, df):
        # Créer une vue calendrier par mois
        mois_examens = {}
        
        for _, examen in df.iterrows():
            date_obj = datetime.strptime(examen['Date'], "%d/%m/%Y")
            mois_annee = date_obj.strftime("%Y-%m")
            
            if mois_annee not in mois_examens:
                mois_examens[mois_annee] = []
            
            mois_examens[mois_annee].append(examen)
        
        for mois, examens in mois_examens.items():
            date_mois = datetime.strptime(mois, "%Y-%m")
            nom_sheet = date_mois.strftime("Examens %B %Y")
            
            # Créer DataFrame pour ce mois
            df_mois = pd.DataFrame(examens)
            df_mois.to_excel(writer, sheet_name=nom_sheet[:31], index=False)  # Limite Excel: 31 caractères
    
    def get_jour_semaine(self, date_obj):
        jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
        return jours[date_obj.weekday()]
    
    def calculer_temps_restant(self, date_examen):
        aujourd_hui = datetime.now().date()
        temps_restant = (date_examen.date() - aujourd_hui).days
        
        if temps_restant < 0:
            return "Passé"
        elif temps_restant == 0:
            return "Aujourd'hui"
        elif temps_restant == 1:
            return "Demain"
        else:
            return f"Dans {temps_restant} jours"
    
    def importer_excel(self):
        fichier = filedialog.askopenfilename(
            title="Sélectionner un fichier Excel",
            filetypes=[("Fichiers Excel", "*.xlsx *.xls")]
        )
        
        if not fichier:
            return
        
        try:
            df = pd.read_excel(fichier)
            
            # Vérifier les colonnes requises
            colonnes_requises = ['Nom de l\'Examen', 'Matière', 'Date', 'Heure de Début', 'Durée (minutes)']
            if not all(col in df.columns for col in colonnes_requises):
                messagebox.showerror("Erreur", "Le fichier Excel ne contient pas toutes les colonnes requises.")
                return
            
            # Importer les examens
            nouveaux_examens = []
            for _, row in df.iterrows():
                try:
                    examen = {
                        'nom': str(row['Nom de l\'Examen']),
                        'matiere': str(row['Matière']),
                        'date': pd.to_datetime(row['Date']).to_pydatetime(),
                        'heure_debut': pd.to_datetime(row['Heure de Début']).time(),
                        'duree': int(row['Durée (minutes)']),
                        'lieu': str(row.get('Lieu', '')),
                        'notes': str(row.get('Notes', ''))
                    }
                    nouveaux_examens.append(examen)
                except Exception as e:
                    continue  # Ignorer les lignes avec des erreurs
            
            self.examens.extend(nouveaux_examens)
            self.mettre_a_jour_liste()
            messagebox.showinfo("Succès", f"{len(nouveaux_examens)} examens importés avec succès!")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'importation: {str(e)}")
    
    def creer_modele_planification(self):
        """Créer un modèle Excel vide avec le format requis pour la planification des examens"""
        try:
            # Créer un nouveau classeur
            wb = Workbook()
            
            # Supprimer la feuille par défaut
            wb.remove(wb.active)
            
            # Définir les styles
            header_font = Font(name='Arial', bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # =============== FEUILLE 1: Salles ===============
            salles_sheet = wb.create_sheet("Salles")
            
            # En-têtes pour la feuille Salles
            salles_headers = ["Nom de la Salle", "Capacité"]
            salles_sheet.append(salles_headers)
            
            # Styliser les en-têtes
            for col_num, header in enumerate(salles_headers, 1):
                cell = salles_sheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Définir la largeur des colonnes
            salles_sheet.column_dimensions['A'].width = 20  # Nom de la Salle
            salles_sheet.column_dimensions['B'].width = 12  # Capacité
            
            # =============== FEUILLE 2: Groupes ===============
            groupes_sheet = wb.create_sheet("Groupes")
            
            # En-têtes pour la feuille Groupes (permettre jusqu'à 5 matières par groupe)
            max_matieres = 5
            groupes_headers = ["Nom du Groupe", "Nb Matières"]
            
            # Ajouter les colonnes Matière et Durée
            for i in range(1, max_matieres + 1):
                groupes_headers.extend([f"Matière {i}", f"Durée {i} (min)"])
            
            groupes_sheet.append(groupes_headers)
            
            # Styliser les en-têtes
            for col_num, header in enumerate(groupes_headers, 1):
                cell = groupes_sheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Définir la largeur des colonnes
            groupes_sheet.column_dimensions['A'].width = 18  # Nom du Groupe
            groupes_sheet.column_dimensions['B'].width = 12  # Nb Matières
            
            # Colonnes Matière et Durée
            for i in range(3, len(groupes_headers) + 1):
                col_letter = groupes_sheet.cell(row=1, column=i).column_letter
                if i % 2 == 1:  # Colonnes Matière
                    groupes_sheet.column_dimensions[col_letter].width = 20
                else:  # Colonnes Durée
                    groupes_sheet.column_dimensions[col_letter].width = 12
            
            # =============== FEUILLE 3: Enseignants ===============
            enseignants_sheet = wb.create_sheet("Enseignants")
            
            # En-têtes pour la feuille Enseignants (permettre jusqu'à 4 matières par enseignant)
            max_matieres_enseignant = 4
            enseignants_headers = ["Nom de l'Enseignant", "Nb Matières"]
            
            for i in range(1, max_matieres_enseignant + 1):
                enseignants_headers.append(f"Matière {i}")
            
            enseignants_sheet.append(enseignants_headers)
            
            # Styliser les en-têtes
            for col_num, header in enumerate(enseignants_headers, 1):
                cell = enseignants_sheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Définir la largeur des colonnes
            enseignants_sheet.column_dimensions['A'].width = 22  # Nom de l'Enseignant
            enseignants_sheet.column_dimensions['B'].width = 12  # Nb Matières
            
            # Colonnes Matière
            for i in range(3, len(enseignants_headers) + 1):
                col_letter = enseignants_sheet.cell(row=1, column=i).column_letter
                enseignants_sheet.column_dimensions[col_letter].width = 20
            
            # =============== AJOUTER FEUILLE INSTRUCTIONS ===============
            instructions_sheet = wb.create_sheet("Instructions")
            
            # Instructions en français
            instructions = [
                ["📋 INSTRUCTIONS - MODE D'EMPLOI"],
                [""],
                ["🏫 FEUILLE 1 - Salles:"],
                ["• Nom de la Salle: Saisissez l'identifiant de la salle (ex: 101, Amphi A)"],
                ["• Capacité: Nombre de groupes que cette salle peut accueillir simultanément"],
                [""],
                ["👥 FEUILLE 2 - Groupes:"],
                ["• Nom du Groupe: Saisissez l'identifiant du groupe (ex: 1ère AP1, Master 1)"],
                ["• Nb Matières: Nombre total de matières pour ce groupe"],
                ["• Matière 1, 2, etc.: Saisissez les noms des matières"],
                ["• Durée 1, 2, etc.: Saisissez la durée de l'examen en minutes"],
                ["• Laissez vides les colonnes Matière/Durée non utilisées"],
                [""],
                ["👨‍🏫 FEUILLE 3 - Enseignants:"],
                ["• Nom de l'Enseignant: Saisissez le nom complet de l'enseignant"],
                ["• Nb Matières: Nombre de matières enseignées par cet enseignant"],
                ["• Matière 1, 2, etc.: Saisissez les matières enseignées"],
                ["• Les enseignants NE PEUVENT PAS surveiller leurs propres matières"],
                ["• Laissez vides les colonnes Matière non utilisées"],
                [""],
                ["⚠️ NOTES IMPORTANTES:"],
                ["• Supprimez cette feuille Instructions avant d'utiliser le planificateur"],
                ["• Toutes les durées doivent être en minutes (ex: 90, 120)"],
                ["• Capacité salle: 1 = groupe unique, 2+ = groupes multiples"],
                ["• Matières enseignant: Utilisées pour éviter l'auto-surveillance"],
                ["• Gardez exactement ces noms de feuilles: Salles, Groupes, Enseignants"],
                [""],
                ["🚀 EXEMPLE D'UTILISATION:"],
                ["Salles: | Salle 101 | 2 | (peut accueillir 2 groupes)"],
                ["Groupes: | Groupe A | 2 | Mathématiques | 120 | Physique | 90 |"],
                ["Enseignants: | Prof. Dupont | 1 | Mathématiques |"]
            ]
            
            # Ajouter les instructions
            for row_num, instruction in enumerate(instructions, 1):
                instructions_sheet.cell(row=row_num, column=1, value=instruction[0])
                
                # Styliser les en-têtes principaux
                if any(word in instruction[0] for word in ["INSTRUCTIONS", "FEUILLE", "NOTES", "EXEMPLE"]):
                    cell = instructions_sheet.cell(row=row_num, column=1)
                    cell.font = Font(bold=True, size=12, color="1f4e79")
            
            # Définir la largeur de colonne pour les instructions
            instructions_sheet.column_dimensions['A'].width = 80
            
            # Déplacer la feuille Instructions en première position
            wb.move_sheet(instructions_sheet, 0)
            
            # Sauvegarder le modèle dans le répertoire de l'exe
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nom_fichier = f"Modele_Planification_Examens_{timestamp}.xlsx"
            chemin_fichier = os.path.join(self.exe_dir, nom_fichier)
            wb.save(chemin_fichier)
            
            messagebox.showinfo("Succès", f"✅ Modèle créé avec succès!\n📁 Fichier: {nom_fichier}\n\n📋 Le modèle inclut:\n• Feuille Instructions (à supprimer avant utilisation)\n• Feuille Salles (prête pour les données)\n• Feuille Groupes (prête pour les données)\n• Feuille Enseignants (prête pour les données)\n\n🚀 Prêt à utiliser avec votre planificateur d'examens!")
            
            # Demander si l'utilisateur veut ouvrir le fichier
            if messagebox.askyesno("Ouvrir le fichier", "Voulez-vous ouvrir le modèle Excel maintenant?"):
                os.startfile(chemin_fichier)
                
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la création du modèle: {str(e)}")

def main():
    root = tk.Tk()
    app = GenerateurCalendrierExamens(root)
    root.mainloop()

if __name__ == "__main__":
    main()