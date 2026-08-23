"""
Convert a "Plan d'étude" workbook (per-specialty sheets, one row per ECUE)
into a "nidham" workbook (one sheet per filière code, UE-level table + ECUE/hours table).

Usage:
    python3 convert_plan.py <input.xlsx> <output.xlsx> <pe_value>

Example:
    python3 convert_plan.py ANCIENS_Plans_..._DESIGN.xlsx nidham_out.xlsx PE.isbat2026
"""
import sys
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# --- sheet name -> (code_suffix, niv_suffix) -------------------------------
# code_suffix is appended directly to the UE code (no separator).
# niv_suffix is appended to the level number to build the 'niv' field.
# Sheets that share the same code_suffix are merged into one output sheet,
# ordered by the semester digit embedded in each UE code.
SHEET_MAPPING = {
    "L1 Design Image S1-S2": ("PG", "PG"),
    "L2 Publicité graphique S3-S4": ("PG", "PG"),
    "L3 Publicité graphique S5-S6": ("PG", "PG"),
    "L2 Art danimation de limage S3-": ("Anim", "Anim"),
    "L3 Art danimation de limage S5-": ("Anim", "Anim"),
    "L1 Design Espace S1-S2": ("AI", "AI"),
    "L2 AI S3-S4": ("AI", "AI"),
    "L3 AI S5-S6": ("AI", "AI"),
    "L2 Scéno S3-S4": ("SC", "Sceno"),
    "L3 Scéno S5-S6": ("SC", "Sceno"),
    "L1 Design Produit S1-S2": ("DP", "DP"),
    "L2 Création Industrielle S3-S4": ("DP", "DP"),
    "L3 Création Industrielle S5-S6": ("DP", "DP"),
}

# code_suffix -> display sheet title (differs for Scéno: code 'SC', tab 'SCENO')
SHEET_TITLES = {"PG": "PG", "Anim": "Anim", "AI": "AI", "SC": "SCENO", "DP": "DP"}

# Arts Plastiques: single track, one sheet per level, all map to code+niv suffix "AP"
AP_SHEET_MAPPING = {
    "L1 ": ("AP", "AP"),
    "L2 ": ("AP", "AP"),
    "L3": ("AP", "AP"),
}
AP_SHEET_TITLES = {"AP": "AP"}

TYPE_MAP = {
    "Obligatoire": "Fond",
    "Transversale": "Trans",
    "Optionnelle": "Opt",
}

# Column indices (0-based) in the source "Plan d'étude" sheets:
# A N° | B UE name | C Code UE | D Type | E ECUE name | F Code ECUE |
# G Cours | H TD | I TP | J Autres | ...
COL_UE_NAME, COL_UE_CODE, COL_TYPE, COL_ECUE_NAME, COL_ECUE_CODE = 1, 2, 3, 4, 5
COL_COURS, COL_TD, COL_TP, COL_AUTRES = 6, 7, 8, 9


def semester_from_code(code_ue: str) -> int:
    """First digit of the numeric part of a UE code gives the semester (UF310 -> 3)."""
    digits = "".join(ch for ch in code_ue if ch.isdigit())
    if not digits:
        raise ValueError(f"Cannot find a semester digit in UE code {code_ue!r}")
    return int(digits[0])


def extract_sheet_rows(ws):
    """
    Walk one source sheet and yield dicts, one per ECUE data row, with the
    UE name/code/type forward-filled across the blank continuation rows that
    the original copy-paste layout leaves behind (merged-cell style input).
    Header rows, title rows and 'Total ...' rows are skipped automatically
    because they have no real ECUE code in column F.
    """
    cur_ue_name = cur_ue_code = cur_type = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        vals = [c.value for c in row]
        if len(vals) <= COL_ECUE_CODE:
            continue
        ue_name, ue_code, ue_type = vals[COL_UE_NAME], vals[COL_UE_CODE], vals[COL_TYPE]
        ecue_name, ecue_code = vals[COL_ECUE_NAME], vals[COL_ECUE_CODE]

        if ue_name and ue_code:
            cur_ue_name, cur_ue_code, cur_type = ue_name, ue_code, ue_type

        # A real data row always has a genuine ECUE code (e.g. 'UF111').
        # Header/title/blank/'Total' rows don't -- skip them.
        if not isinstance(ecue_code, str) or not any(ch.isdigit() for ch in ecue_code):
            continue
        if cur_ue_code is None:
            continue  # stray data before any UE header seen -- shouldn't happen

        cours = vals[COL_COURS] if len(vals) > COL_COURS else None
        td = vals[COL_TD] if len(vals) > COL_TD else None
        tp = vals[COL_TP] if len(vals) > COL_TP else None
        autres = vals[COL_AUTRES] if len(vals) > COL_AUTRES else None

        # Observed rule (from the AP/design benchmark): when hours live only in
        # 'Autres' (Cours/TD/TP all empty), split them 50/50 into Cours and TD.
        if cours is None and td is None and tp is None and autres is not None:
            cours = td = autres / 2
            autres = None

        yield {
            "ue_name": cur_ue_name,
            "ue_code": cur_ue_code,
            "ue_type": cur_type,
            "ecue_name": ecue_name,
            "ecue_code": ecue_code,
            "cours": cours,
            "td": td,
            "tp": tp,
            "autres": autres,
        }


def build_filiere_data(wb, code_suffix, mapping):
    """Collect all rows for every source sheet mapped to this code_suffix."""
    rows = []
    for sheet_name, (suf, niv_suf) in mapping.items():
        if suf != code_suffix or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in extract_sheet_rows(ws):
            r["niv_suffix"] = niv_suf
            r["sem"] = semester_from_code(r["ue_code"])
            rows.append(r)
    rows.sort(key=lambda r: r["sem"])  # stable: keeps original N° order within a semester
    return rows


def write_filiere_sheet(wb_out, code_suffix, sheet_title, rows, pe_value):
    ws = wb_out.create_sheet(title=sheet_title)

    header_font = Font(name="Arial", bold=True)
    header = ["Unité d'enseignement (UE)", "Code de l'UE",
              "Type de l'UE\n(Obligatoire / Optionnelle)", "sem", "pe", "niv"]
    for i, h in enumerate(header, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = header_font

    # ---- Table 1: one row per UE (per semester occurrence) ----
    seen_ue = set()
    r = 3
    for row in rows:
        key = row["ue_code"]
        if key in seen_ue:
            continue
        seen_ue.add(key)
        niv = f"{(row['sem'] + 1) // 2}{row['niv_suffix']}"
        ws.cell(row=r, column=1, value=row["ue_name"])
        ws.cell(row=r, column=2, value=f"{row['ue_code']}{code_suffix}")
        ws.cell(row=r, column=3, value=TYPE_MAP.get(row["ue_type"], row["ue_type"]))
        ws.cell(row=r, column=4, value=float(row["sem"]))
        ws.cell(row=r, column=5, value=pe_value)
        ws.cell(row=r, column=6, value=niv)
        r += 1

    r += 2  # blank rows before the ECUE/hours table

    # ---- Table 2: one row per ECUE, hours per 14 weeks + weekly (/14) ----
    ecue_header = ["Code de l'UE", "Elément constitutif d'UE (ECUE)", "Code de l'ECUE",
                   "Cours", "TD", "TP", "Autres", "Cours", "TD", "TP", "Autres"]
    for i, h in enumerate(ecue_header, start=1):
        ws.cell(row=r, column=i, value=h)
    r += 1
    for row in rows:
        ws.cell(row=r, column=1, value=f"{row['ue_code']}{code_suffix}")
        ws.cell(row=r, column=2, value=row["ecue_name"])
        ws.cell(row=r, column=3, value=row["ecue_code"])
        ws.cell(row=r, column=4, value=row["cours"])
        ws.cell(row=r, column=5, value=row["td"])
        ws.cell(row=r, column=6, value=row["tp"])
        ws.cell(row=r, column=7, value=row["autres"])
        for offset, col in enumerate("DEFG", start=0):
            ws.cell(row=r, column=8 + offset, value=f"={col}{r}/14")
        r += 1

    return ws


def convert(input_path, output_path, pe_value, mapping=SHEET_MAPPING, sheet_titles=SHEET_TITLES):
    wb_in = openpyxl.load_workbook(input_path, data_only=False)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    code_suffixes = []
    for suf, _ in mapping.values():
        if suf not in code_suffixes:
            code_suffixes.append(suf)

    for suf in code_suffixes:
        rows = build_filiere_data(wb_in, suf, mapping)
        if rows:
            write_filiere_sheet(wb_out, suf, sheet_titles.get(suf, suf), rows, pe_value)

    wb_out.save(output_path)
    print(f"Wrote {output_path} with sheets: {wb_out.sheetnames}")


if __name__ == "__main__":
    if len(sys.argv) > 4 and sys.argv[4] == "ap":
        convert(sys.argv[1], sys.argv[2], sys.argv[3], AP_SHEET_MAPPING, AP_SHEET_TITLES)
    else:
        convert(sys.argv[1], sys.argv[2], sys.argv[3])
