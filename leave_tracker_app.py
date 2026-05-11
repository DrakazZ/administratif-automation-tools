# leave_tracker_app.py
import sys
import os
import sqlite3
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QComboBox, QMessageBox, QTabWidget, QSpinBox, QDateEdit
)
from PyQt5.QtCore import QDate
from PyQt5.QtGui import QIcon

def get_resource_path(relative_path):
    # Resolve files for both dev runs and PyInstaller onefile builds.
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

def get_app_dir():
    # Use AppData to ensure persistence in one-file EXE builds.
    base_dir = os.getenv("APPDATA") or os.path.expanduser("~")
    app_dir = os.path.join(base_dir, "LeaveTracker")
    os.makedirs(app_dir, exist_ok=True)
    return app_dir

APP_DIR = get_app_dir()
DB_PATH = os.path.join(APP_DIR, "leave_data.db")
APP_ICON_PNG = get_resource_path("isbat.png")

def get_export_dir():
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    export_dir = os.path.join(desktop_dir, "Rapports Conges")
    os.makedirs(export_dir, exist_ok=True)
    return export_dir

EXPORT_DIR = get_export_dir()

def get_export_path(filename):
    return os.path.join(EXPORT_DIR, filename)

def save_workbook(workbook, filename):
    export_path = get_export_path(filename)
    try:
        workbook.save(export_path)
        return export_path, False
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        name, ext = os.path.splitext(filename)
        fallback_name = f"{name}_{timestamp}{ext}"
        fallback_path = get_export_path(fallback_name)
        workbook.save(fallback_path)
        return fallback_path, True

LEAVE_TYPES = {
    'سنوية': {'yearly_add': 45, 'reset': False},
    'استثنائية': {'yearly_add': 6, 'reset': True},
    'مرضية': {'yearly_add': 0, 'reset': False},
    'تعوضية': {'yearly_add': 0, 'reset': False},
    'اضافية': {'yearly_add': 0, 'reset': False},
}

def get_column_index(cursor, column_name):
    for i, col in enumerate(cursor.description):
        if col[0] == column_name:
            return i
    raise ValueError(f"Column '{column_name}' not found in results")

def get_table_columns(cursor, table_name):
    cursor.execute(f"PRAGMA table_info({table_name})")
    return {row[1] for row in cursor.fetchall()}

class LeaveApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("منضومة العطل")
        self.resize(600, 400)
        self.setWindowIcon(QIcon(APP_ICON_PNG))
        self.create_db()

        self.tabs = QTabWidget()
        self.tabs.addTab(self.add_employee_tab(), "اضافة موظف")
        self.tabs.addTab(self.manage_leave_tab(), "ادارة العطل")
        self.tabs.addTab(self.summary_tab(), "مخلص العطل")
        self.tabs.addTab(self.reset_tab(),"اعادة ضبط سنوي")

        layout = QVBoxLayout()
        layout.addWidget(self.tabs)
        self.setLayout(layout)

    def create_db(self):
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        fields = ", ".join([
            f"{t}_taken INTEGER DEFAULT 0, {t}_left INTEGER DEFAULT {LEAVE_TYPES[t]['yearly_add']}"
            for t in LEAVE_TYPES
        ])
        cursor.execute(f"""
            CREATE TABLE IF NOT EXISTS employees (
                reg_number TEXT PRIMARY KEY,
                name TEXT,
                surname TEXT,
                rank TEXT,
                {fields}
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS leave_history (
                reg_number TEXT,
                year INTEGER,
                leave_type TEXT,
                taken INTEGER,
                PRIMARY KEY (reg_number, year, leave_type)
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS leave_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                reg_number TEXT,
                year INTEGER,
                leave_type TEXT,
                days INTEGER,
                start_date TEXT,
                end_date TEXT,
                action TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS reset_log (
                year INTEGER PRIMARY KEY,
                performed_at TEXT DEFAULT (datetime('now'))
            )
        """)

        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_leave_events_reg_year
            ON leave_events (reg_number, year)
        """)

        employee_columns = get_table_columns(cursor, "employees")
        if "rank" not in employee_columns:
            cursor.execute("ALTER TABLE employees ADD COLUMN rank TEXT DEFAULT ''")

        conn.commit()
        conn.close()

    def add_employee_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        self.name_input = QLineEdit()
        self.surname_input = QLineEdit()
        self.rank_input = QLineEdit()
        self.reg_input = QLineEdit()
        layout.addWidget(QLabel("الاسم"))
        layout.addWidget(self.name_input)
        layout.addWidget(QLabel("اللقب"))
        layout.addWidget(self.surname_input)
        layout.addWidget(QLabel("الرتبة"))
        layout.addWidget(self.rank_input)
        layout.addWidget(QLabel("المعرف الوحيد"))
        layout.addWidget(self.reg_input)

        self.initial_left_inputs = {}
        for t in LEAVE_TYPES:
            input_field = QSpinBox()
            input_field.setRange(0, 365)
            self.initial_left_inputs[t] = input_field
            layout.addWidget(QLabel(f"  عدد ايام العطل {t} المتبقية"))
            layout.addWidget(input_field)

        add_btn = QPushButton("اضافة موظف")
        add_btn.clicked.connect(self.add_employee)
        layout.addWidget(add_btn)

        tab.setLayout(layout)
        return tab

    def manage_leave_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        self.reg_manage = QLineEdit()
        self.days_input = QSpinBox()
        self.days_input.setRange(1, 365)
        self.leave_type = QComboBox()
        self.leave_type.addItems(LEAVE_TYPES.keys())

        self.start_date_input = QDateEdit()
        self.start_date_input.setCalendarPopup(True)
        self.start_date_input.setDate(QDate.currentDate())

        self.end_date_input = QDateEdit()
        self.end_date_input.setCalendarPopup(True)
        self.end_date_input.setReadOnly(True)
        self.end_date_input.setDate(QDate.currentDate())

        layout.addWidget(QLabel("المعرف الوحيد"))
        layout.addWidget(self.reg_manage)
        layout.addWidget(QLabel("عدد ايام العطل"))
        layout.addWidget(self.days_input)
        layout.addWidget(QLabel("نوع العطل"))
        layout.addWidget(self.leave_type)
        layout.addWidget(QLabel("تاريخ البداية"))
        layout.addWidget(self.start_date_input)
        layout.addWidget(QLabel("تاريخ النهاية"))
        layout.addWidget(self.end_date_input)

        btn_layout = QHBoxLayout()
        take_btn = QPushButton("اخذ ايام عطل")
        take_btn.clicked.connect(self.take_days)
        add_btn = QPushButton("اضافة ايام عطل")
        add_btn.clicked.connect(self.add_days)
        btn_layout.addWidget(take_btn)
        btn_layout.addWidget(add_btn)

        layout.addLayout(btn_layout)
        tab.setLayout(layout)

        self.days_input.valueChanged.connect(self.update_end_date)
        self.start_date_input.dateChanged.connect(self.update_end_date)
        self.update_end_date()
        return tab

    def summary_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        self.reg_summary = QLineEdit()
        self.year_input = QSpinBox()
        self.year_input.setRange(2000, 2100)
        self.year_input.setValue(datetime.now().year)

        self.summary_label = QLabel("")

        layout.addWidget(QLabel("المعرف الوحيد"))
        layout.addWidget(self.reg_summary)
        layout.addWidget(QLabel("السنة"))
        layout.addWidget(self.year_input)

        summary_btn = QPushButton("عرض الملخص")
        summary_btn.clicked.connect(self.show_summary)
        layout.addWidget(summary_btn)

        # ✅ Export personal summary to Excel
        export_btn = QPushButton("تحميل ملخص الموظف (Excel)")
        export_btn.clicked.connect(self.export_selected_summary)
        layout.addWidget(export_btn)

        # ✅ Export all summaries to Excel
        export_all_btn = QPushButton("تحميل ملخص كل الموظفين (Excel)")
        export_all_btn.clicked.connect(self.export_all_summaries)
        layout.addWidget(export_all_btn)

        layout.addWidget(self.summary_label)
        tab.setLayout(layout)
        return tab


    def reset_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        reset_btn = QPushButton("اعادة ضبط سنوي")
        reset_btn.clicked.connect(self.perform_reset)
        layout.addWidget(reset_btn)

        tab.setLayout(layout)
        return tab

    def add_employee(self):
        name = self.name_input.text()
        surname = self.surname_input.text()
        rank = self.rank_input.text()
        reg = self.reg_input.text()

        if not (name and surname and rank and reg):
            QMessageBox.warning(self, "معلومات خاطئة", "الرجاء ادخال المعلومات المطلوبة")
            return

        values = {
            f"{t}_left": self.initial_left_inputs[t].value() for t in LEAVE_TYPES
        }

        columns = ", ".join(["reg_number", "name", "surname", "rank"] + list(values.keys()))
        placeholders = ", ".join(["?"] * len(values))
        sql = f"INSERT INTO employees ({columns}) VALUES (?, ?, ?, ?, {placeholders})"

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute(sql, (reg, name, surname, rank, *values.values()))
            conn.commit()
            QMessageBox.information(self, "نجاح", "تم اضافة الموظف بنجاح")
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "خطا", "المعرف الوحيد مستخدم من قبل")
        conn.close()

    def take_days(self):
        self.modify_days(-1)

    def add_days(self):
        self.modify_days(1)

    def update_end_date(self):
        start_qdate = self.start_date_input.date()
        days = self.days_input.value()
        start = date(start_qdate.year(), start_qdate.month(), start_qdate.day())
        end = start + timedelta(days=days - 1)
        self.end_date_input.setDate(QDate(end.year, end.month, end.day))

    def split_leave_by_year(self, start_date, end_date):
        segments = []
        current_start = start_date

        while current_start.year < end_date.year:
            current_end = date(current_start.year, 12, 31)
            segment_days = (current_end - current_start).days + 1
            segments.append((current_start, current_end, current_start.year, segment_days))
            current_start = date(current_start.year + 1, 1, 1)

        segment_days = (end_date - current_start).days + 1
        segments.append((current_start, end_date, current_start.year, segment_days))
        return segments


    def modify_days(self, direction):
        reg = self.reg_manage.text()
        days = self.days_input.value()
        leave_type = self.leave_type.currentText()
        start_qdate = self.start_date_input.date()
        start_date = date(start_qdate.year(), start_qdate.month(), start_qdate.day())
        end_date = start_date + timedelta(days=days - 1)

        if not reg or not leave_type or days <= 0:
            QMessageBox.warning(self, "خطأ", "الرجاء ملء كل الحقول بشكل صحيح")
            return

        conn = sqlite3.connect(DB_PATH)
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM employees WHERE reg_number = ?", (reg,))
            row = cursor.fetchone()

            if not row:
                QMessageBox.warning(self, "خطأ", "الموظف غير موجود")
                return

            taken_col = f"{leave_type}_taken"
            left_col = f"{leave_type}_left"
            col_names = [col[0] for col in cursor.description]

            try:
                left_index = col_names.index(left_col)
                taken_index = col_names.index(taken_col)
            except ValueError:
                QMessageBox.warning(self, "خطأ", "نوع العطلة غير موجود في قاعدة البيانات")
                return

            current_left = row[left_index]

            if direction == -1:
                # ✅ Taking leave — restrict if not enough days
                if current_left < days:
                    QMessageBox.warning(self, "أيام غير كافية", "لا يوجد أيام كافية لهذه العطلة")
                    return

                # Update employee leave balance
                cursor.execute(f"""
                    UPDATE employees SET
                        {taken_col} = {taken_col} + ?,
                        {left_col} = {left_col} - ?
                    WHERE reg_number = ?
                """, (days, days, reg))

                segments = self.split_leave_by_year(start_date, end_date)
                for seg_start, seg_end, seg_year, seg_days in segments:
                    cursor.execute("""
                        INSERT INTO leave_events
                            (reg_number, year, leave_type, days, start_date, end_date, action)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (
                        reg,
                        seg_year,
                        leave_type,
                        seg_days,
                        seg_start.isoformat(),
                        seg_end.isoformat(),
                        "take",
                    ))


            elif direction == 1:
                # ✅ Adding days — allow resulting negative or high values (no restriction)
                cursor.execute(f"""
                    UPDATE employees SET
                        {left_col} = {left_col} + ?
                    WHERE reg_number = ?
                """, (days, reg))

                cursor.execute("""
                    INSERT INTO leave_events
                        (reg_number, year, leave_type, days, start_date, end_date, action)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    reg,
                    start_date.year,
                    leave_type,
                    days,
                    start_date.isoformat(),
                    end_date.isoformat(),
                    "adjust_add",
                ))

            conn.commit()
        finally:
            conn.close()
        QMessageBox.information(self, "تحديث", "تم تحديث بيانات العطلة بنجاح")

    def show_summary(self):
        reg = self.reg_summary.text()
        year = self.year_input.value()
        conn = sqlite3.connect(DB_PATH)
        try:
            cursor = conn.cursor()

            # Fetch employee data first
            cursor.execute("SELECT * FROM employees WHERE reg_number = ?", (reg,))
            row = cursor.fetchone()
            if not row:
                QMessageBox.warning(self, "خطأ", "الموظف غير موجود")
                return

            employee_cols = [col[0] for col in cursor.description]
            rank_index = employee_cols.index("rank") if "rank" in employee_cols else None
            name, surname = row[1], row[2]
            rank = row[rank_index] if rank_index is not None else ""
            summary = f"الاسم: {name} {surname}\nالرتبة: {rank}\nالسنة: {year}\n\n"

            cursor.execute("""
                SELECT leave_type, SUM(days)
                FROM leave_events
                WHERE reg_number = ? AND year = ? AND action = ?
                GROUP BY leave_type
            """, (reg, year, "take"))
            history = {row[0]: row[1] for row in cursor.fetchall()}

            for t in LEAVE_TYPES:
                taken = history.get(t, 0)
                try:
                    left_index = employee_cols.index(f"{t}_left")
                    left = row[left_index]
                except ValueError:
                    left = "؟"

                summary += f"{t}: أخذ = {taken}, متبقي حاليا = {left}\n"

            self.summary_label.setText(summary)
        finally:
            conn.close()



    def perform_reset(self):
        year = datetime.now().year
        today = datetime.now()
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        RESET_ALLOWED_DAY = 10

        # Check if today is before Jan 10th
        if today.month == 1 and today.day <= RESET_ALLOWED_DAY:
            QMessageBox.warning(self, "تحذير", f"الرجاء الانتظار حتى {RESET_ALLOWED_DAY} يناير للقيام بإعادة الضبط السنوي.")
            conn.close()
            return

        cursor.execute("SELECT 1 FROM reset_log WHERE year = ?", (year,))
        if cursor.fetchone():
            QMessageBox.warning(self, "تحذير", f"تم بالفعل إعادة الضبط لهذه السنة: {year}")
            conn.close()
            return

        reply = QMessageBox.question(self, "تأكيد", f"هل أنت متأكد أنك تريد إعادة ضبط السنة {year}؟",
                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            conn.close()
            return

        backup_filename = f"نسخة_احتياطية_قبل_إعادة_الضبط_{year}.xlsx"
        self.export_current_summary_to_excel(backup_filename)

        cursor.execute("INSERT INTO reset_log (year, performed_at) VALUES (?, ?)", (
            year,
            datetime.now().isoformat(timespec="seconds"),
        ))

        for t, props in LEAVE_TYPES.items():
            if props['reset']:
                cursor.execute(f"UPDATE employees SET {t}_left = {props['yearly_add']}, {t}_taken = 0")
            else:
                cursor.execute(f"UPDATE employees SET {t}_left = {t}_left + {props['yearly_add']}, {t}_taken = 0")

        conn.commit()
        conn.close()
        QMessageBox.information(self, "تمت العملية", "تم اعادة الضبط السنوي العطل بنجاح")

    def export_employee_summary_to_excel(self, reg, year):
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Get employee data and store column info
        cursor.execute("SELECT * FROM employees WHERE reg_number = ?", (reg,))
        employee_row = cursor.fetchone()
        if not employee_row:
            QMessageBox.warning(self, "خطأ", "الموظف غير موجود")
            conn.close()
            return

        # Store column names from employee table BEFORE changing queries
        employee_cols = [col[0] for col in cursor.description]
        rank_index = employee_cols.index("rank") if "rank" in employee_cols else None
        name, surname = employee_row[1], employee_row[2]
        rank = employee_row[rank_index] if rank_index is not None else ""
        
        cursor.execute("""
            SELECT leave_type, SUM(days)
            FROM leave_events
            WHERE reg_number = ? AND year = ? AND action = ?
            GROUP BY leave_type
        """, (reg, year, "take"))
        history = {row[0]: row[1] for row in cursor.fetchall()}

        wb = Workbook()
        ws = wb.active
        ws.title = f"ملخص {reg}"

        ws.append(["رقم التسجيل", "الاسم", "اللقب", "الرتبة", "السنة"])
        ws.append([reg, name, surname, rank, year])
        ws.append([])
        ws.append(["نوع العطلة", "الأيام المستعملة", "الأيام المتبقية (حاليا)"])

        for t in LEAVE_TYPES:
            taken = history.get(t, 0)
            # Use employee_cols instead of cursor.description
            try:
                left_index = employee_cols.index(f"{t}_left")
                left = employee_row[left_index]
            except ValueError:
                left = "غير متوفر"  # Fallback if column doesn't exist
            
            ws.append([t, taken, left])

        details_ws = wb.create_sheet(title="تفاصيل العطل")
        details_ws.append(["نوع العطلة", "عدد الايام", "تاريخ البداية", "تاريخ النهاية"])
        cursor.execute("""
            SELECT leave_type, days, start_date, end_date
            FROM leave_events
            WHERE reg_number = ? AND year = ? AND action = ?
            ORDER BY start_date
        """, (reg, year, "take"))
        for row in cursor.fetchall():
            details_ws.append(list(row))

        filename = f"ملخص_{reg}_{year}.xlsx"
        saved_path, used_fallback = save_workbook(wb, filename)
        conn.close()
        if used_fallback:
            QMessageBox.warning(self, "تم الحفظ", f"تعذر الكتابة على الملف المطلوب. تم حفظ نسخة جديدة في:\n{saved_path}")
        else:
            QMessageBox.information(self, "تم الحفظ", f"تم حفظ ملخص الموظف في الملف:\n{saved_path}")

    
    def export_current_summary_to_excel(self, filename="نسخة_احتياطية_قبل_إعادة_الضبط.xlsx"):
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        wb = Workbook()
        ws = wb.active
        ws.title = "نسخة احتياطية"

        headers = ["رقم التسجيل", "الاسم", "اللقب"] + [
            f"المستعملة ({t})" for t in LEAVE_TYPES
        ] + [
            f"المتبقية ({t})" for t in LEAVE_TYPES
        ]
        headers.insert(3, "الرتبة")
        ws.append(headers)

        cursor.execute("SELECT * FROM employees")
        employee_cols = [col[0] for col in cursor.description]
        rank_index = employee_cols.index("rank") if "rank" in employee_cols else None
        rows = cursor.fetchall()
        for row in rows:
            rank = row[rank_index] if rank_index is not None else ""
            base = [row[0], row[1], row[2], rank]
            taken = [row[get_column_index(cursor, f"{t}_taken")] for t in LEAVE_TYPES]
            left = [row[get_column_index(cursor, f"{t}_left")] for t in LEAVE_TYPES]
            ws.append(base + taken + left)

        saved_path, used_fallback = save_workbook(wb, filename)
        conn.close()

        if used_fallback:
            QMessageBox.warning(self, "تم الحفظ", f"تعذر الكتابة على الملف المطلوب. تم حفظ نسخة جديدة في:\n{saved_path}")
        else:
            QMessageBox.information(self, "تم الحفظ", f"تم حفظ الملخص في الملف:\n{saved_path}")

    def export_all_summaries_to_excel(self, year):
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        wb = Workbook()
        ws = wb.active
        ws.title = f"ملخص {year}"

        headers = ["رقم التسجيل", "الاسم", "اللقب", "الرتبة"] + [
            f"المستعملة ({t})" for t in LEAVE_TYPES
        ] + [
            f"المتبقية ({t})" for t in LEAVE_TYPES
        ]
        ws.append(headers)

        # Get all employee data and store column info
        cursor.execute("SELECT * FROM employees")
        all_employee_rows = cursor.fetchall()
        employee_cols = [col[0] for col in cursor.description]  # Store column names
        rank_index = employee_cols.index("rank") if "rank" in employee_cols else None

        for employee_row in all_employee_rows:
            reg = employee_row[0]
            name = employee_row[1]
            surname = employee_row[2]
            rank = employee_row[rank_index] if rank_index is not None else ""
            
            cursor.execute("""
                SELECT leave_type, SUM(days)
                FROM leave_events
                WHERE reg_number = ? AND year = ? AND action = ?
                GROUP BY leave_type
            """, (reg, year, "take"))
            history = {row[0]: row[1] for row in cursor.fetchall()}
            
            # Build taken and left arrays using stored column info
            taken = [history.get(t, 0) for t in LEAVE_TYPES]
            left = []
            
            for t in LEAVE_TYPES:
                try:
                    left_index = employee_cols.index(f"{t}_left")
                    left.append(employee_row[left_index])
                except ValueError:
                    left.append("غير متوفر")  # Fallback if column doesn't exist
            
            ws.append([reg, name, surname, rank] + taken + left)

        filename = f"ملخص_جميع_الموظفين_{year}.xlsx"
        saved_path, used_fallback = save_workbook(wb, filename)
        conn.close()
        if used_fallback:
            QMessageBox.warning(self, "تم الحفظ", f"تعذر الكتابة على الملف المطلوب. تم حفظ نسخة جديدة في:\n{saved_path}")
        else:
            QMessageBox.information(self, "تم الحفظ", f"تم حفظ الملخص في الملف:\n{saved_path}")

    def export_selected_summary(self):
        reg = self.reg_summary.text()
        year = self.year_input.value()
        if not reg:
            QMessageBox.warning(self, "خطأ", "يرجى إدخال رقم التسجيل أولاً")
            return
        self.export_employee_summary_to_excel(reg, year)

    def export_all_summaries(self):
        year = self.year_input.value()
        self.export_all_summaries_to_excel(year)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(APP_ICON_PNG))
    window = LeaveApp()
    window.show()
    sys.exit(app.exec_())
